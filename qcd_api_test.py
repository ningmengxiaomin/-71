"""
python自动化测试实现前准备
①接口测试用例
②python代码读取接口测试用例
③requests库发送请求
④执行结果 VS 预期结果----测试用例是否通过
⑤响应结果的回写
"""
import openpyxl
import requests

def read_qcd_case(filename,worksheet):
    "读取测试用例的数据"
    case_file = openpyxl.load_workbook(filename)                #加载excel表
    case_sheet = case_file[worksheet]                           #获取excel表中的名字为：register的sheet
    row_max = case_sheet.max_row                                #获取表格中的最大行
    case_list = []                                              #把每个案例的url,data,expected的字典存放在一个列表中
    for i in range(2,row_max+1):#用for遍历表格中的每一行数据
        case_id = case_sheet.cell(row=i,column=1).value
        data_url = case_sheet.cell(row=i,column=5).value        #获取表格中每行中的Url数据
        data_par = case_sheet.cell(row=i,column=6).value        #获取表格中每行中的data数据
        data_expected = case_sheet.cell(row=i,column=7).value   #获取表格中每行中的expected数据
        case_dict = dict(id=case_id,url=data_url,data=data_par,expected=data_expected)  #将每行的Url、data、expected的数据以字典的形式进行存储
        case_list.append(case_dict)                             #把每行的Url、data、expected的字典存放在大列表中
    #print(case_list)
    return case_list

def wirte_result(filename, worksheet,rows,columns,case_result):
    "向测试用例中写入测试结果"
    case_filename = openpyxl.load_workbook(filename)          #加载工作薄--测试用例的excel表
    case_sheet = case_filename[worksheet]                     #获取sheet
    case_sheet.cell(row=rows,column=columns).value=case_result    #在实际结果的列中写入测试的结果
    case_filename.save(filename)        #写入结果后进行保存

def qcd_api_request(filename, worksheet, token):
    "接口测试用例的执行—requests发送请求"
    list_case = read_qcd_case(filename, worksheet)      #调用读取测试用例的函数，获取返回值
    for i in range(len(list_case)):                     #使用for循环遍历list_data数据列表
        case_id = list_case[i]["id"]
        url = list_case[i]["url"]                       #通过字典key来获取url数据
        data_par = eval(list_case[i]["data"])           #通过字典key来取data数据，使用eval（）函数来脱掉python包裹代码的引号
        expected = list_case[i]["expected"]             #通过字典key来取预期结果
        expected = eval(expected)                       #使用eval（）函数来脱掉python包裹代码的引号
        expected_msg = expected.get("msg")              #通过字典key来取预期结果中的msg的数据
        respones = requests.post(url,json=data_par,headers=token)   #发送接口请求
        #print(respones.json())      #以json数据格式输出响应结果
        result = respones.json()["msg"]                 #获取响应结果中的msg的数据
        if result == expected_msg:
            print("第{}条测试用例执行通过" .format(case_id))
            case_result = "Pass"
        else:
            print("第{}条测试用例执行不通过" .format(case_id))
            case_result = "Fail"
        wirte_result(filename, worksheet,case_id+1,8,case_result)  #调用写入测试结果的函数

token={"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
qcd_api_request("test_case_api.xlsx","register",token)




